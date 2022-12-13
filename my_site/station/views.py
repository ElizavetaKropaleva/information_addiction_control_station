# Основные функции, обеспечивающие работу сайта

from django.http import HttpResponse
from django.shortcuts import render, redirect, reverse
from django.core.files.storage import FileSystemStorage
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.contrib import auth

from .models import Slot_1, Slot_2, Slot_3, Slot_4, Slot_5, Slot_6
import os


# Create your views here.
# Функция управления домашней странцей
def home(request):

    # Отбор актуальной информации из таблиц БД и передача этих данных на домашнюю страницу
    n_1 = Slot_1.objects.all().count()
    n_2 = Slot_2.objects.all().count()
    n_3 = Slot_3.objects.all().count()
    n_4 = Slot_4.objects.all().count()
    n_5 = Slot_5.objects.all().count()
    n_6 = Slot_6.objects.all().count()
    info = [Slot_1.objects.get(pk=n_1), Slot_2.objects.get(pk=n_2), Slot_3.objects.get(pk=n_3),
            Slot_4.objects.get(pk=n_4), Slot_5.objects.get(pk=n_5), Slot_6.objects.get(pk=n_6)]
    return render(request, 'home.html',
                  {'info_1': info[0], 'info_2': info[1], 'info_3': info[2], 'info_4': info[3], 'info_5': info[4],
                   'info_6': info[5], 'user': auth.get_user(request).username})


# Обработка полученных данных из формы
def loading_data(request):
    if request.method == 'POST' and request.FILES['data_file']:
        data_file = request.FILES['data_file']
        fs = FileSystemStorage()
        filename = fs.save(data_file.name, data_file)
        uploaded_file_url = fs.url(filename)
        parser('C:\Python\PD\information_addiction_control_station' + uploaded_file_url)
        return render(request, 'loading_data.html', {'uploaded_file_url': uploaded_file_url, 'user': auth.get_user(request).username})
    return render(request, 'loading_data.html', {'user': auth.get_user(request).username})


# Дополнение данных в слотах
def data_addition(request):
    if request.POST:
        # получение данных из формы
        slot = request.POST['inputSlot']
        inputName = request.POST['inputName']
        inputModel = request.POST['inputModel']
        # внесение изменений в соответствии с номером слота
        if slot == "1":
            n = Slot_1.objects.all().count()
            Slot_1.objects.filter(id=n).update(name=inputName, device_model=inputModel)
        if slot == "2":
            n = Slot_2.objects.all().count()
            Slot_2.objects.filter(id=n).update(name=inputName, device_model=inputModel)
        if slot == "3":
            n = Slot_3.objects.all().count()
            Slot_3.objects.filter(id=n).update(name=inputName, device_model=inputModel)
        if slot == "4":
            n = Slot_4.objects.all().count()
            Slot_4.objects.filter(id=n).update(name=inputName, device_model=inputModel)
        if slot == "5":
            n = Slot_5.objects.all().count()
            Slot_5.objects.filter(id=n).update(name=inputName, device_model=inputModel)
        if slot == "6":
            n = Slot_6.objects.all().count()
            Slot_6.objects.filter(id=n).update(name=inputName, device_model=inputModel)
        return redirect(reverse('station:home'))
    else:
        return render(request, 'data_addition.html', {'user': auth.get_user(request).username})


# Получение статистики
def statistics(request):
    if request.method == 'POST':
        slot = request.POST['inputSlot']
        firstDate = request.POST['inputFirstDate']
        secondDate = request.POST['inputSecondDate']

        # Получение необходимых данных с выбранного слота
        if slot == "1":
            data = Slot_1.objects.filter(date__gte=firstDate) & Slot_1.objects.filter(date__lte=secondDate)
            title = "Слот 1"
        if slot == "2":
            data = Slot_2.objects.filter(date__gte=firstDate) & Slot_2.objects.filter(date__lte=secondDate)
            title = "Слот 2"
        if slot == "3":
            data = Slot_3.objects.filter(date__gte=firstDate) & Slot_3.objects.filter(date__lte=secondDate)
            title = "Слот 3"
        if slot == "4":
            data = Slot_4.objects.filter(date__gte=firstDate) & Slot_4.objects.filter(date__lte=secondDate)
            title = "Слот 4"
        if slot == "5":
            data = Slot_5.objects.filter(date__gte=firstDate) & Slot_5.objects.filter(date__lte=secondDate)
            title = "Слот 5"
        if slot == "6":
            data = Slot_6.objects.filter(date__gte=firstDate) & Slot_6.objects.filter(date__lte=secondDate)
            title = "Слот 6"

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        # Создание Excel-файла
        workbook = Workbook()
        # Получение активного листа в файле
        worksheet = workbook.active
        worksheet.title = title
        # Настройка названия столбцов
        columns = [
            (' Дата', 18),
            (' Имя владельца', 18),
            (' Модель устройства', 22),
            (' Уровень заряда', 18),
            (' Время подключения', 25),
            (' Время отключения', 25)
        ]
        row_num = 1
        # Настройка заголовков для каждой ячейки
        for col_num, (column_title, column_width) in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            # установка ширины
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = column_width
        # Внесение данных
        for i in data:
            row_num += 1
            row = [
                i.date,
                i.name,
                i.device_model,
                i.charge,
                i.connection_time,
                i.disconnection_time
            ]
            # Настройка данных для каждой ячейки строки
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        # Сохранение excel-файла
        workbook.save(response)
        return response
    return render(request, 'statistics.html', {'user': auth.get_user(request).username})


# Страница для авторизации
def login(request):
    # Отбор актуальной информации из таблиц БД и передача этих данных на домашнюю страницу
    n_1 = Slot_1.objects.all().count()
    n_2 = Slot_2.objects.all().count()
    n_3 = Slot_3.objects.all().count()
    n_4 = Slot_4.objects.all().count()
    n_5 = Slot_5.objects.all().count()
    n_6 = Slot_6.objects.all().count()
    info = [Slot_1.objects.get(pk=n_1), Slot_2.objects.get(pk=n_2), Slot_3.objects.get(pk=n_3),
            Slot_4.objects.get(pk=n_4), Slot_5.objects.get(pk=n_5), Slot_6.objects.get(pk=n_6)]

    args = {}
    if request.POST:
        username = request.POST['username']
        password = request.POST['password']
        user = auth.authenticate(username=username, password=password)
        # проверка подлинности
        if user is not None:
            auth.login(request, user) # успешная авторизация
            return render(request, 'home.html',
                          {'info_1': info[0], 'info_2': info[1], 'info_3': info[2], 'info_4': info[3],
                           'info_5': info[4],
                           'info_6': info[5], 'user': auth.get_user(request).username})
        else:
            args['login_error'] = "Пользователь не найден" # сообщение о неудачной авторизации
            return render(request, "login.html", args)
    else:
        return render(request, "login.html", args)


# Функция выхода из аккаунта
def logout(request):
    # Отбор актуальной информации из таблиц БД и передача этих данных на домашнюю страницу
    n_1 = Slot_1.objects.all().count()
    n_2 = Slot_2.objects.all().count()
    n_3 = Slot_3.objects.all().count()
    n_4 = Slot_4.objects.all().count()
    n_5 = Slot_5.objects.all().count()
    n_6 = Slot_6.objects.all().count()
    info = [Slot_1.objects.get(pk=n_1), Slot_2.objects.get(pk=n_2), Slot_3.objects.get(pk=n_3),
            Slot_4.objects.get(pk=n_4), Slot_5.objects.get(pk=n_5), Slot_6.objects.get(pk=n_6)]

    auth.logout(request) # разлогирование
    return render(request, 'home.html',
                  {'info_1': info[0], 'info_2': info[1], 'info_3': info[2], 'info_4': info[3],
                   'info_5': info[4],
                   'info_6': info[5], 'user': auth.get_user(request).username})


# Функция парсера
def parser(file):
    # Работа с файлом
    f = open(file)
    flag = True
    while flag:
        # Считываем файл построчно и разделяем данные по пробелу в массив
        line = f.readline()
        if not line:
            flag = False
            break
        arr = line.split()

        # Анализируем номер слота (первые данные в строке)
        if arr[0] == '1':
            # Заполняем значениям по умолчанию
            name = 'name'
            device_model = 'device model'
            # Заполняем значениями из массива
            charge = arr[1]
            connection_time = arr[2]
            # Анализируем наличие устройства в слоте по данным о времени отсоединения
            if len(arr) == 4:
                disconnection_time = arr[3]
            else:
                disconnection_time = 'non'
            # Создаем новую запись в таблице
            Slot_1.objects.create(name=name, device_model=device_model, charge=charge, connection_time=connection_time,
                                  disconnection_time=disconnection_time, date=datetime.now().date())

        # Анализируем номер слота (первые данные в строке)
        if arr[0] == '2':
            # Заполняем значениям по умолчанию
            name = 'name'
            device_model = 'device model'
            # Заполняем значениями из массива
            charge = arr[1]
            connection_time = arr[2]
            # Анализируем наличие устройства в слоте по данным о времени отсоединения
            if len(arr) == 4:
                disconnection_time = arr[3]
            else:
                disconnection_time = 'non'
            # Создаем новую запись в таблице
            Slot_2.objects.create(name=name, device_model=device_model, charge=charge, connection_time=connection_time,
                                  disconnection_time=disconnection_time, date=datetime.now().date())

        # Анализируем номер слота (первые данные в строке)
        if arr[0] == '3':
            # Заполняем значениям по умолчанию
            name = 'name'
            device_model = 'device model'
            # Заполняем значениями из массива
            charge = arr[1]
            connection_time = arr[2]
            # Анализируем наличие устройства в слоте по данным о времени отсоединения
            if len(arr) == 4:
                disconnection_time = arr[3]
            else:
                disconnection_time = 'non'
            # Создаем новую запись в таблице
            Slot_3.objects.create(name=name, device_model=device_model, charge=charge, connection_time=connection_time,
                                  disconnection_time=disconnection_time, date=datetime.now().date())

        # Анализируем номер слота (первые данные в строке)
        if arr[0] == '4':
            # Заполняем значениям по умолчанию
            name = 'name'
            device_model = 'device model'
            # Заполняем значениями из массива
            charge = arr[1]
            connection_time = arr[2]
            # Анализируем наличие устройства в слоте по данным о времени отсоединения
            if len(arr) == 4:
                disconnection_time = arr[3]
            else:
                disconnection_time = 'non'
            # Создаем новую запись в таблице
            Slot_4.objects.create(name=name, device_model=device_model, charge=charge, connection_time=connection_time,
                                  disconnection_time=disconnection_time, date=datetime.now().date())

        # Анализируем номер слота (первые данные в строке)
        if arr[0] == '5':
            # Заполняем значениям по умолчанию
            name = 'name'
            device_model = 'device model'
            # Заполняем значениями из массива
            charge = arr[1]
            connection_time = arr[2]
            # Анализируем наличие устройства в слоте по данным о времени отсоединения
            if len(arr) == 4:
                disconnection_time = arr[3]
            else:
                disconnection_time = 'non'
            # Создаем новую запись в таблице
            Slot_5.objects.create(name=name, device_model=device_model, charge=charge, connection_time=connection_time,
                                  disconnection_time=disconnection_time, date=datetime.now().date())

        # Анализируем номер слота (первые данные в строке)
        if arr[0] == '6':
            # Заполняем значениям по умолчанию
            name = 'name'
            device_model = 'device model'
            # Заполняем значениями из массива
            charge = arr[1]
            connection_time = arr[2]
            # Анализируем наличие устройства в слоте по данным о времени отсоединения
            if len(arr) == 4:
                disconnection_time = arr[3]
            else:
                disconnection_time = 'non'
            # Создаем новую запись в таблице
            Slot_6.objects.create(name=name, device_model=device_model, charge=charge, connection_time=connection_time,
                                  disconnection_time=disconnection_time, date=datetime.now().date())

    f.close()
    os.remove(file)  # удаление файла из памяти
